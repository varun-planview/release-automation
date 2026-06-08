from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path


@dataclass
class Card:
    card_id: str
    team: str
    card_title: str
    assignees_raw: str
    card_type: str
    lane_title: str
    planning_increment_label: str
    tags: str
    pap_id: str = ""
    card_url: str = ""
    assignee_names: str = ""
    release_versions: list[str] = field(default_factory=list)


def _parse_assignees(raw: str) -> str:
    """Split 'Last, First; Last2, First2' into 'First, First2'."""
    if not raw:
        return ""
    names = []
    for part in raw.split(";"):
        part = part.strip()
        if not part:
            continue
        sub = [s.strip() for s in part.split(",")]
        # PS1 behavior: takes index [1] (first name after comma)
        names.append(sub[1] if len(sub) >= 2 else sub[0])
    return ", ".join(names)


def read_cards(path: Path, base_url: str) -> list[Card]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("ERROR: openpyxl not installed. Run: pip install openpyxl")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers: list[str] = []
    cards: list[Card] = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx == 0:
            headers = [str(c).strip() if c is not None else "" for c in row]
            continue

        row_data = dict(zip(headers, row))
        raw_id = str(row_data.get("Card ID", "") or "").strip()
        # Mirror PS1: keep rows where Card ID contains at least one digit
        if not re.search(r"\d+", raw_id):
            continue

        pil = str(row_data.get("Planning Increment Label", "") or "")
        release_versions = [v.strip() for v in pil.split(",") if v.strip()] if pil else []

        assignees_raw = str(row_data.get("Assignee(s)", "") or "")
        card = Card(
            card_id=raw_id,
            team=str(row_data.get("Team", "") or ""),
            card_title=str(row_data.get("Card Title", "") or ""),
            assignees_raw=assignees_raw,
            card_type=str(row_data.get("Card Type", "") or ""),
            lane_title=str(row_data.get("Current Lane Title", "") or ""),
            planning_increment_label=pil,
            tags=str(row_data.get("Tags", "") or ""),
            pap_id=f"PAP-{raw_id}",
            card_url=f"{base_url}/{raw_id}",
            assignee_names=_parse_assignees(assignees_raw),
            release_versions=release_versions,
        )
        cards.append(card)

    wb.close()
    return cards
